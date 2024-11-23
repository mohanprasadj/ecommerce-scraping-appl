import requests
from bs4 import BeautifulSoup
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import datetime
import shutil
import os
import time
from pathlib import Path

#Pre-Requisites
#Drop box account with the dropbox folder in your PC
#Drop box folder containing product_list.xlsx with the products whose data you want to get often
#Example folder available in this repository

MAX_NUM_PAGE = 4 # No of pages
RUN_FREQUENCY = 6 # In hours
DELETION_TIME = 86400 # 1 Day
REFRESH_TIME = 10 # In seconds
CATEGORIES = []
AMAZON_DROPBOX_DIRECTORY = "/home/test/Dropbox/repository/amazon_product" #Dropbox directory where product data will be saved
AMAZON_PRODUCT_LIST = "/home/test/Dropbox/repository/product_list.xlsx" #File which contains list of products whose data we want to get in a frequency RUN_FREQUENCY
results_by_category = {}

def get_product_list(platform):
    global CATEGORIES
    if not os.path.exists(AMAZON_PRODUCT_LIST):
        return False
    df = pd.read_excel(AMAZON_PRODUCT_LIST, sheet_name=platform)
    CATEGORIES = df.iloc[:, 0].tolist()
    return True

def scrape_amazon_products(search_term, num_pages=1):
    headers = {
        'User-Agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:129.0) Gecko/20100100 Firefox/129.0',
        'Accept-Language': 'en-US, en;q=0.5'
    }
    
    results = []
    
    for page in range(1, num_pages + 1):
        url = f"https://www.amazon.in/s?k={search_term}&page={page}"
        response = requests.get(url, headers=headers)
        
        if response.status_code != 200:
            print(f"Failed to retrieve page {page} for {search_term}")
            time.sleep(REFRESH_TIME)
            continue
        
        soup = BeautifulSoup(response.content, 'html.parser')
        
        products = soup.find_all('div', {'data-component-type': 's-search-result'})
        
        for product in products:
            title_element = product.find('span', class_='a-size-medium a-color-base a-text-normal')
            price_element = product.find('span', class_='a-price-whole')
            original_price_element = product.find_all('span', class_='a-offscreen')
            original_price_element = original_price_element[1] if len(original_price_element) > 1 else None

            title = title_element.text.strip() if title_element else 'N/A'
            price = price_element.text.strip() if price_element else 'N/A'
            original_price = original_price_element.text.strip() if original_price_element else 'N/A'
 
            try:
                price = int(price.replace(',', '')) if price != 'N/A' else 0
            except Exception:
                price = 0
            match = re.search(r'\d{1,3}(?:,\d{3})*', original_price)
            orig_price = int(match.group(0).replace(',', '')) if match else 0
            
            discount = int(((orig_price - price) / orig_price * 100)) if orig_price > 0 and price > 0 else 0

            results.append({
                'Title': title,
                'Price': price,
                'Original Price': orig_price,
                'Discount': discount
            })
        time.sleep(REFRESH_TIME)
    
    return results

def save_to_excel(data_dict, filename='amazon_products.xlsx'):
    workbook = Workbook()
    
    for category_name, data in data_dict.items():
        sheet = workbook.create_sheet(title=category_name[:31])  # Limit sheet name to 31 characters
        
        # Write headers
        headers = ['Title', 'Price', 'Original Price', 'Discount']
        sheet.append(headers)
        
        # Write data and apply conditional formatting
        for row in data:
            sheet.append([row['Title'], row['Price'], row['Original Price'], row['Discount']])
        
        # Apply conditional formatting to the Discount column (4th column)
        for row in range(2, len(data) + 2):  # Start from row 2 to skip header
            discount_cell = sheet.cell(row=row, column=4)  # Column D (Discount)
            
            # Yellow fill for discounts between 50 and 80
            if 50 <= discount_cell.value < 80:
                discount_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            # Green fill for discounts between 80 and 100
            elif 80 <= discount_cell.value <= 100:
                discount_cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    # Remove the default sheet created with the workbook
    if "Sheet" in workbook.sheetnames:
        workbook.remove(workbook["Sheet"])

    # Save the workbook
    workbook.save(filename)

def delete_old_files(directory):
    for root, dirs, files in os.walk(directory):
        for file in files:
            file_path = os.path.join(root, file)
            # Get the last modified time of the file
            file_mod_time = os.path.getmtime(file_path)
            # Check if the file is older than 1 day
            if time.time() - file_mod_time > DELETION_TIME:
                os.remove(file_path)

def main():
    already_run = False
    while True:
        if datetime.datetime.now().hour % RUN_FREQUENCY == 0 and already_run is False:
            print(f"Starting to fetch at {datetime.datetime.now()}")
            got_list = get_product_list("amazon")
            if not got_list:
                already_run = True
                continue
            delete_old_files(AMAZON_DROPBOX_DIRECTORY)
            for category in CATEGORIES:
                products_data = scrape_amazon_products(category, num_pages=MAX_NUM_PAGE)
                results_by_category[category] = products_data
            formatted_date = datetime.datetime.now().strftime('%d-%m-%y')
            formatted_time = datetime.datetime.now().strftime('%H-%S-%M')
            file = f"amazon_products_{formatted_date}_{formatted_time}.xlsx"
            save_to_excel(results_by_category, filename = file)
            shutil.copy(os.path.join(os.getcwd(), file), AMAZON_DROPBOX_DIRECTORY)
            print(f"Report saved to {file}")
            already_run = True
        else:
            if datetime.datetime.now().hour % RUN_FREQUENCY != 0:
                already_run = False

if __name__ == "__main__":
    main()
